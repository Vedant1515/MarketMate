import logging
from datetime import date
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel, Field

from app.services.ingestion import run_ingestion
from app.services.sales_store import SalesStoreError, get_sales_store
from app.services.vector_store import get_vector_store
from app.tools.demand_forecaster import forecast_item, invalidate_cache


logger = logging.getLogger(__name__)

router = APIRouter()


class DailySaleRecord(BaseModel):
    item: str
    quantity_sold: float = Field(gt=0)
    unit_price_aud: float = Field(gt=0)
    date: Optional[str] = None
    is_public_holiday: int = 0
    is_pre_holiday: int = 0


class DailySalesRequest(BaseModel):
    records: List[DailySaleRecord]
    date: Optional[str] = None


@router.get("/api/sales/items")
async def get_items() -> List[Dict[str, Any]]:
    try:
        store = get_sales_store()
        return store.get_items()
    except Exception as exc:
        logger.error("GET /api/sales/items error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/api/sales/daily")
async def log_daily_sales(request: DailySalesRequest) -> Dict[str, Any]:
    log_date = request.date or date.today().strftime("%Y-%m-%d")

    records = []
    for rec in request.records:
        records.append({
            "item": rec.item,
            "quantity_sold": rec.quantity_sold,
            "unit_price_aud": rec.unit_price_aud,
            "date": rec.date or log_date,
            "is_public_holiday": rec.is_public_holiday,
            "is_pre_holiday": rec.is_pre_holiday,
        })

    try:
        store = get_sales_store()
        added = store.append_daily_sales(records)
    except Exception as exc:
        logger.error("POST /api/sales/daily store error: %s", exc)
        raise HTTPException(status_code=500, detail=f"Failed to save sales data: {exc}")

    # Invalidate demand forecaster cache so it reads fresh data
    invalidate_cache()

    # Re-ingest into ChromaDB so RAG reflects the new data
    ingested = 0
    if added > 0:
        try:
            vs = get_vector_store()
            settings_path = store._path
            ingested = run_ingestion(vs, settings_path)
            logger.info("Re-ingested %d documents after sales log update", ingested)
        except Exception as exc:
            logger.warning("Re-ingestion after sales log failed (data saved, RAG not updated): %s", exc)

    weeks = store.get_weeks_of_data()

    return {
        "rows_added": added,
        "documents_reindexed": ingested,
        "total_weeks_of_data": weeks,
        "message": (
            f"Saved {added} sales records. ChromaDB updated with {ingested} documents. "
            f"System now has {weeks} weeks of data."
        ) if added > 0 else "No new records added (duplicates skipped).",
    }


@router.get("/api/sales/forecast/{item_name}")
async def get_forecast(item_name: str) -> Dict[str, Any]:
    try:
        invalidate_cache()
        result = forecast_item(item_name)
        if "error" in result:
            raise HTTPException(status_code=404, detail=result["error"])
        return result
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("GET /api/sales/forecast/%s error: %s", item_name, exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/api/sales/stats")
async def get_sales_stats() -> Dict[str, Any]:
    try:
        store = get_sales_store()
        vs = get_vector_store()
        return {
            "weeks_of_data": store.get_weeks_of_data(),
            "latest_week": store.get_latest_week_number(),
            "documents_in_chromadb": vs.document_count(),
            "items": len(store.get_items()),
        }
    except Exception as exc:
        logger.error("GET /api/sales/stats error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/api/sales/template")
async def download_template(format: str = "csv") -> Response:
    store = get_sales_store()
    csv_content = store.generate_template_csv()

    if format == "csv":
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=marketmate_sales_template.csv"},
        )

    # Generate xlsx
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Sales"

        items = store.get_items()
        item_names = [i["item"] for i in items]

        header = ["Date"] + item_names
        ws.append(header)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1a7a3f")
        for col_idx, _ in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        # Add two example rows
        from datetime import date as dt, timedelta
        today = dt.today()
        for offset in range(1, 3):
            row_date = (today - timedelta(days=offset)).strftime("%Y-%m-%d")
            ws.append([row_date] + [0] * len(item_names))

        # Set column widths
        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, len(header) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 14

        # Add instructions sheet
        ws2 = wb.create_sheet("Instructions")
        ws2["A1"] = "MarketMate Sales Upload Template"
        ws2["A1"].font = Font(bold=True, size=13)
        instructions = [
            "",
            "How to fill in the Daily Sales sheet:",
            "1. Column A (Date): enter each day in YYYY-MM-DD or DD/MM/YYYY format.",
            "2. Columns B onwards: enter the total quantity sold for that item on that day.",
            "3. Leave as 0 if an item was not sold.",
            "4. Each row = one day. You can include multiple days.",
            "5. Upload the file using the 'Upload Excel' tab in MarketMate.",
            "",
            "Supported formats: .xlsx, .xls, .csv",
            "The AI updates its knowledge base automatically after each upload.",
        ]
        for row_text in instructions:
            ws2.append([row_text])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=marketmate_sales_template.xlsx"},
        )
    except Exception as exc:
        logger.error("Template xlsx generation failed: %s", exc)
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=marketmate_sales_template.csv"},
        )


@router.post("/api/sales/upload")
async def upload_sales_file(file: UploadFile = File(...)) -> Dict[str, Any]:
    allowed = (".csv", ".xlsx", ".xls")
    filename = file.filename or "upload"
    ext = "." + filename.lower().rsplit(".", 1)[-1]

    if ext not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Allowed: {', '.join(allowed)}",
        )

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum 5MB.")

    store = get_sales_store()

    try:
        records = store.parse_upload_file(content, filename)
    except SalesStoreError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    except Exception as exc:
        logger.error("File parse error for '%s': %s", filename, exc, exc_info=True)
        raise HTTPException(status_code=422, detail=f"Failed to parse file: {exc}")

    if not records:
        raise HTTPException(
            status_code=422,
            detail="No valid sales records found in the file. Check the format matches the template.",
        )

    try:
        added = store.append_daily_sales(records)
    except Exception as exc:
        logger.error("Upload append failed: %s", exc)
        raise HTTPException(status_code=500, detail=f"Failed to save records: {exc}")

    invalidate_cache()

    ingested = 0
    if added > 0:
        try:
            vs = get_vector_store()
            ingested = run_ingestion(vs, store._path)
        except Exception as exc:
            logger.warning("Re-ingestion after upload failed: %s", exc)

    unique_dates = list({r["date"] for r in records})
    unique_dates.sort()

    return {
        "filename": filename,
        "records_parsed": len(records),
        "rows_added": added,
        "documents_reindexed": ingested,
        "total_weeks_of_data": store.get_weeks_of_data(),
        "dates_found": unique_dates,
        "items_found": list({r["item"] for r in records}),
        "message": (
            f"Imported {added} records from {len(unique_dates)} day(s). "
            f"ChromaDB updated with {ingested} documents. "
            f"System now has {store.get_weeks_of_data()} weeks of data."
        ) if added > 0 else f"Parsed {len(records)} records but all were duplicates.",
    }
